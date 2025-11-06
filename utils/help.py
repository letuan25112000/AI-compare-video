import shutil
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import subprocess
from pathlib import Path
import tempfile
import os
import cv2
import json
from PIL import Image, ImageDraw, ImageFont
import numpy as np

from config import FFMPEG_PATH

# =====================================
# フォルダ内の古いファイルを削除
# =====================================
def clean_folder(results_dir):
    """
    指定フォルダ内のファイル・サブフォルダを全削除
    """
    results_dir = Path(results_dir)
    for f in results_dir.glob("*"):
        try:
            if f.is_file() or f.is_symlink():
                f.unlink()
            elif f.is_dir():
                shutil.rmtree(f)
        except Exception as e:
            print(f"古いファイル削除エラー: {f} → {e}")


# =====================================
# タイムスタンプ付きファイル名生成
# =====================================
def timestamp_file_path(filename, ext):
    """
    ファイル名に現在時刻（YYYYMMDD_HHMMSS）を付与して返す
    例: result_20251023_153000.xlsx
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{filename}_{timestamp}.{ext}"


# =====================================
# Excel書式設定
# =====================================
def apply_excel_format(ws):
    """
    Excel表のフォーマットを適用：
    ヘッダーのスタイル、枠線、セルの整列、背景色、自動列幅調整など
    """

    # --- ヘッダー行の書式 ---
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    center_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- 枠線（細線）設定 ---
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- 全セルに枠線と整列を適用 ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row != 1:  # ヘッダーは既に中央揃え済み
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- 列幅を自動調整 ---
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                width = len(str(cell.value)) + 2
                column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 10), width)

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


# =====================================
# Web再生対応の動画へ変換（ffmpeg使用）
# =====================================
def make_web_ready(input_path):
    """
    OpenCVで出力したMP4ファイルを
    ブラウザで再生できる形式（H.264 + yuv420p）に変換する。
    """
    input_path = Path(input_path)
    output_path = input_path.with_name(input_path.stem + "_web.mp4")

    cmd = [
        FFMPEG_PATH,
        "-y",                # 既存ファイルを上書き
        "-i", str(input_path),
        "-vcodec", "libx264",
        "-pix_fmt", "yuv420p",    # ブラウザ再生に必要
        "-movflags", "+faststart",  # ストリーミング高速開始
        str(output_path)
    ]

    subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return str(output_path)

# =====================================
# 動画のfps同期
# =====================================
def convert_to_nfps(video_path, fps=30):
    """
    入力動画を指定FPSに変換して一時ファイルとして保存する
    """
    temp_dir = Path(tempfile.gettempdir())
    output_path = temp_dir / f"{Path(video_path).stem}_{fps}fps.mp4"

    # ffmpeg コマンド
    cmd = [
        FFMPEG_PATH, "-y",
        "-i", str(video_path),
        "-filter:v", f"fps={fps}",
        "-c:v", "libx264",
        "-preset", "ultrafast",
        "-an",  # 音声を削除
        str(output_path)
    ]

    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    return str(output_path)


# =====================================
# 新しい機能: 動画のメタデータ取得
# =====================================
def get_video_metadata(video_path):
    """
    動画のメタデータを取得（解像度、FPS、長さなど）
    """
    try:
        cap = cv2.VideoCapture(str(video_path))
        if not cap.isOpened():
            return None
            
        fps = cap.get(cv2.CAP_PROP_FPS)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps if fps > 0 else 0
        
        cap.release()
        
        return {
            'fps': fps,
            'width': width,
            'height': height,
            'frame_count': frame_count,
            'duration': duration,
            'resolution': f"{width}x{height}"
        }
    except Exception as e:
        print(f"メタデータ取得エラー: {e}")
        return None


# =====================================
# 新しい機能: フレーム比較画像の生成
# =====================================
def create_comparison_image(org_frame_path, comp_frame_path, output_path, diff_info=None):
    """
    2つのフレームを横に並べて比較画像を生成
    """
    try:
        # 画像を読み込み
        org_img = Image.open(org_frame_path)
        comp_img = Image.open(comp_frame_path)
        
        # 同じサイズにリサイズ
        width = max(org_img.width, comp_img.width)
        height = max(org_img.height, comp_img.height)
        
        org_img = org_img.resize((width, height))
        comp_img = comp_img.resize((width, height))
        
        # 比較画像を作成（横に並べる）
        comparison = Image.new('RGB', (width * 2, height), (255, 255, 255))
        comparison.paste(org_img, (0, 0))
        comparison.paste(comp_img, (width, 0))
        
        # 差分情報を追加（オプション）
        if diff_info:
            draw = ImageDraw.Draw(comparison)
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except:
                font = ImageFont.load_default()
            
            # タイトルと差分情報
            title = "Frame Comparison"
            added_text = f"Added: {', '.join(diff_info.get('added', []))}" if diff_info.get('added') else ""
            removed_text = f"Removed: {', '.join(diff_info.get('removed', []))}" if diff_info.get('removed') else ""
            
            # テキスト描画
            draw.text((10, 10), title, fill=(0, 0, 0), font=font)
            draw.text((10, 40), f"Original", fill=(0, 0, 255), font=font)
            draw.text((width + 10, 40), f"Comparison", fill=(255, 0, 0), font=font)
            
            y_pos = 70
            if added_text:
                draw.text((10, y_pos), added_text, fill=(0, 128, 0), font=font)
                y_pos += 25
            if removed_text:
                draw.text((10, y_pos), removed_text, fill=(255, 0, 0), font=font)
        
        comparison.save(output_path)
        return True
    except Exception as e:
        print(f"比較画像生成エラー: {e}")
        return False


# =====================================
# 新しい機能: 動画のサムネイル生成
# =====================================
def generate_video_thumbnails(video_path, output_dir, num_thumbnails=5):
    """
    動画から複数のサムネイルを生成
    """
    try:
        cap = cv2.VideoCapture(str(video_path))
        if not cap.isOpened():
            return []
        
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        interval = max(frame_count // num_thumbnails, 1)
        
        thumbnails = []
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
        
        for i in range(num_thumbnails):
            frame_pos = i * interval
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_pos)
            ret, frame = cap.read()
            
            if ret:
                thumbnail_path = output_dir / f"thumb_{i+1}_{Path(video_path).stem}.jpg"
                cv2.imwrite(str(thumbnail_path), frame)
                thumbnails.append(str(thumbnail_path))
        
        cap.release()
        return thumbnails
    except Exception as e:
        print(f"サムネイル生成エラー: {e}")
        return []


# =====================================
# 新しい機能: 動画のトリミング
# =====================================
def trim_video(video_path, output_path, start_time=0, end_time=None):
    """
    動画を指定時間でトリミング
    """
    try:
        cmd = [
            FFMPEG_PATH, "-y",
            "-i", str(video_path),
            "-ss", str(start_time),
        ]
        
        if end_time:
            cmd.extend(["-to", str(end_time)])
        
        cmd.extend([
            "-c", "copy",
            str(output_path)
        ])
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            return True
        else:
            print(f"トリミングエラー: {result.stderr}")
            return False
    except Exception as e:
        print(f"トリミング処理エラー: {e}")
        return False


# =====================================
# 新しい機能: 動画の結合
# =====================================
def concatenate_videos(video_list, output_path):
    """
    複数の動画を結合
    """
    try:
        # 一時ファイルに動画リストを作成
        list_file = Path(tempfile.gettempdir()) / "video_list.txt"
        with open(list_file, 'w', encoding='utf-8') as f:
            for video in video_list:
                f.write(f"file '{Path(video).absolute()}'\n")
        
        cmd = [
            FFMPEG_PATH, "-y",
            "-f", "concat",
            "-safe", "0",
            "-i", str(list_file),
            "-c", "copy",
            str(output_path)
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        list_file.unlink(missing_ok=True)  # 一時ファイル削除
        
        if result.returncode == 0:
            return True
        else:
            print(f"動画結合エラー: {result.stderr}")
            return False
    except Exception as e:
        print(f"動画結合処理エラー: {e}")
        return False


# =====================================
# 新しい機能: 動画情報のJSONエクスポート
# =====================================
def export_video_info_to_json(video_paths, output_path):
    """
    動画ファイルの情報をJSONにエクスポート
    """
    video_info = {}
    
    for video_path in video_paths:
        metadata = get_video_metadata(video_path)
        if metadata:
            video_info[Path(video_path).name] = {
                'file_path': str(video_path),
                'metadata': metadata,
                'file_size': Path(video_path).stat().st_size,
                'modified_time': Path(video_path).stat().st_mtime
            }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(video_info, f, indent=2, ensure_ascii=False)
    
    return len(video_info)


# =====================================
# 新しい機能: 画像の品質圧縮
# =====================================
def compress_image(input_path, output_path, quality=85):
    """
    画像を圧縮して保存
    """
    try:
        img = Image.open(input_path)
        if img.mode in ('RGBA', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[-1])
            img = background
        
        img.save(output_path, 'JPEG', quality=quality, optimize=True)
        return True
    except Exception as e:
        print(f"画像圧縮エラー: {e}")
        return False


# =====================================
# 新しい機能: 動画からGIF生成
# =====================================
def create_gif_from_video(video_path, output_path, duration=5, fps=10):
    """
    動画からGIFを生成
    """
    try:
        cmd = [
            FFMPEG_PATH, "-y",
            "-i", str(video_path),
            "-t", str(duration),
            "-vf", f"fps={fps},scale=320:-1:flags=lanczos",
            "-c:v", "gif",
            str(output_path)
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        return result.returncode == 0
    except Exception as e:
        print(f"GIF生成エラー: {e}")
        return False


# =====================================
# 新しい機能: バッチ処理用ユーティリティ
# =====================================
def batch_process_videos(input_dir, output_dir, process_function, file_extensions=('.mp4', '.avi', '.mov')):
    """
    ディレクトリ内の動画をバッチ処理
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    processed_files = []
    
    for ext in file_extensions:
        for video_file in input_dir.glob(f"*{ext}"):
            try:
                output_file = output_dir / f"processed_{video_file.name}"
                if process_function(str(video_file), str(output_file)):
                    processed_files.append(str(output_file))
            except Exception as e:
                print(f"バッチ処理エラー {video_file}: {e}")
    
    return processed_files


# =====================================
# 新しい機能: ファイルサイズフォーマット
# =====================================
def format_file_size(size_bytes):
    """
    ファイルサイズを人間が読みやすい形式に変換
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.2f} {size_names[i]}"


# =====================================
# 新しい機能: 進捗表示ユーティリティ
# =====================================
class ProgressTracker:
    """
    処理の進捗を追跡するクラス
    """
    def __init__(self, total_steps, description="Processing"):
        self.total_steps = total_steps
        self.current_step = 0
        self.description = description
    
    def update(self, step=1, message=""):
        self.current_step += step
        progress = (self.current_step / self.total_steps) * 100
        print(f"\r{self.description}: {progress:.1f}% ({self.current_step}/{self.total_steps}) {message}", end="")
        if self.current_step >= self.total_steps:
            print()  # 完了時に改行


# =====================================
# 新しい機能: 動画のハッシュ計算（重複検出用）
# =====================================
def calculate_video_hash(video_path, sample_frames=10):
    """
    動画のハッシュ値を計算（重複検出用）
    """
    try:
        cap = cv2.VideoCapture(str(video_path))
        if not cap.isOpened():
            return None
        
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        interval = max(frame_count // sample_frames, 1)
        
        hash_values = []
        for i in range(sample_frames):
            frame_pos = i * interval
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_pos)
            ret, frame = cap.read()
            
            if ret:
                # フレームをリサイズしてハッシュ計算
                small_frame = cv2.resize(frame, (16, 16))
                gray_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2GRAY)
                frame_hash = hash(gray_frame.tobytes())
                hash_values.append(frame_hash)
        
        cap.release()
        
        # 全フレームのハッシュから動画のハッシュを計算
        return hash(tuple(hash_values))
    except Exception as e:
        print(f"ハッシュ計算エラー: {e}")
        return None